import logging
from datetime import datetime
import gspread
import openpyxl

from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Text
from apscheduler.schedulers.asyncio import AsyncIOScheduler

import config as cfg
import keyboards as kb
import utils as fsm
from db import DataBase

logging.basicConfig(level=logging.INFO)

storage = MemoryStorage()
bot = Bot(token=cfg.ADMIN_BOT_TOKEN, parse_mode='HTML')
user_bot = Bot(token=cfg.USER_BOT_TOKEN, parse_mode='HTML')
dp = Dispatcher(bot=bot, storage=storage)
db = DataBase('bodhi.db')

deleted_msgs = list()


async def delete_msgs(chat_id) -> None:
    already_deleted_msg = list()
    for msg in deleted_msgs:
        if msg['chat_id'] == chat_id:
            try:
                await bot.delete_message(chat_id=msg['chat_id'], message_id=msg['message_id'])
                already_deleted_msg.append(msg)
            except Exception as ex:
                print(ex)
    for msg in already_deleted_msg:
        deleted_msgs.remove(msg)
    already_deleted_msg.clear()


async def statistic(admin_id, username, state: FSMContext) -> None:
    user_id = db.get_user_id(username)[0]
    print(user_id)
    questions = db.get_questions()
    statistic_list = db.get_user_statisctic(user_id)
    answers = list()
    dates = list()
    for question in questions:
        answer_list = list()
        date_list = list()
        for i in statistic_list:
            if i[1] == question[1]:
                date_list.append(i[0])
                answer_list.append(i[2])
        answers.append(answer_list)
        dates.append(date_list)
    title_row = list()
    title_row.append('Вопрос')
    title_row.extend(dates[0])

    book = openpyxl.Workbook()
    book.remove(book.active)
    sh = book.create_sheet('Статистика')
    sh.append(title_row)
    sh.column_dimensions['A'].width = 50
    for i in range(len(questions)):
        row = list()
        row.append(questions[i][1])
        row.extend(answers[i])
        sh.append(row)

    filename = 'statistic/' + username + '.xlsx'
    book.save(filename)
    file = open(filename, 'rb')
    await bot.send_document(admin_id, file)
    await state.finish()


async def all_statistics(admin_id) -> None:
    await bot.send_message(chat_id=admin_id, text='Идет генерация файла статистики')
    admin = db.get_admin(admin_id)
    role = admin[5]
    users = set()
    if role == 3:
        users = db.get_users()
    elif role == 2:
        users = db.get_users_curator(admin_id)
    else:
        print('Нет прав на статистику')
    questions = db.get_questions()

    book = openpyxl.Workbook()
    book.remove(book.active)
    for user in users:
        statistic_list = db.get_user_statisctic(user[1])
        answers = list()
        dates = list()
        for question in questions:
            answer_list = list()
            date_list = list()
            for i in statistic_list:
                if i[1] == question[1]:
                    date_list.append(i[0])
                    answer_list.append(i[2])
            answers.append(answer_list)
            dates.append(date_list)
        title_row = list()
        title_row.append('Вопрос')
        title_row.extend(dates[0])
        sh = book.create_sheet(user[2])
        sh.append(title_row)
        sh.column_dimensions['A'].width = 50
        for i in range(len(questions)):
            row = list()
            row.append(questions[i][1])
            row.extend(answers[i])
            sh.append(row)

    filename = 'statistic/Ответы_о_практике_' + str(admin_id) + '.xlsx'
    book.save(filename)
    file = open(filename, 'rb')
    await bot.send_document(admin_id, file)

    did_train_info = db.get_did_train()
    dates = list()
    users_ids = list()
    users_usernames = list()
    for i in did_train_info:
        if i[3] not in dates:
            dates.append(i[3])
        if i[0] not in users_ids:
            users_ids.append(i[0])
            users_usernames.append(i[1])
    did_train_statistic = list()
    for i in range(len(users_ids)):
        row = list()
        row.append(users_usernames[i])
        for date in dates:
            check = 0
            for info in did_train_info:
                if info[0] == users_ids[i] and info[3] == date and check == 0:
                    if info[2] == 1:
                        row.append('+')
                    elif info[2] == 0:
                        row.append('-')
                    check = 1
            if not check:
                row.append('нет ответа')
        did_train_statistic.append(row)
    title_row = list()
    title_row.append('Username')
    title_row.extend(dates)
    did_train_book = openpyxl.Workbook()
    did_train_book.remove(did_train_book.active)
    did_train_sh = did_train_book.create_sheet('Тренировка')
    did_train_sh.append(title_row)
    did_train_sh.column_dimensions['A'].width = 30
    for i in did_train_statistic:
        did_train_sh.append(i)
    filename = 'statistic/Кто_выполнил_тренировку_' + str(admin_id) + '.xlsx'
    did_train_book.save(filename)
    file = open(filename, 'rb')
    await bot.send_document(admin_id, file)


@dp.message_handler(state=fsm.Statistic.username)
async def username_statistic(msg: types.Message, state: FSMContext) -> None:
    username = msg.text.replace('@', '')
    if db.user_exists_username(username):
        await statistic(msg.from_user.id, username, state)
    else:
        await msg.answer('Пользователь не найден!')


# Проверка функционала
@dp.message_handler(commands=['check'], state='*')
async def cmd_check(msg: types.Message, state: FSMContext) -> None:
    await bot.send_audio(chat_id=msg.from_user.id, audio=types.InputFile('files/Новая запись 84.m4a'),
                         caption='text', reply_markup=kb.one_btn('ГО', 'ГО'))
    # await bot.send_sticker(chat_id=msg.from_user.id, sticker='CAACAgIAAxkBAAMDZO-aoLd36-WmbVVya00LN8JXfmsAAng3AAL57nlL2HqayX6XLVMwBA')


@dp.message_handler(commands=['stop'], state='*')
async def stop_state(msg: types.Message, state: FSMContext) -> None:
    current_state = await state.get_state()
    if current_state is not None:
        await state.finish()
        await msg.answer('Статус остановлен!')
    else:
        await msg.answer('Статут был пустым!')


# Получить ID стикера
@dp.message_handler(content_types=types.ContentType.STICKER, state='*')
async def sticker_id(msg: types.Message, state: FSMContext) -> None:
    file_id = msg.sticker.file_id
    await msg.answer(file_id)
    await bot.send_sticker(chat_id=msg.from_user.id, sticker=file_id)
    await state.update_data(file_id=file_id)
    await fsm.Sticker.code.set()


@dp.message_handler(state=fsm.Sticker.code)
async def sticker_id(msg: types.Message, state: FSMContext) -> None:
    code = msg.text
    data = await state.get_data()
    sticker_id = db.add_sticker(code, data['file_id'])
    await msg.answer(f'Стикер добавлен {sticker_id}')
    await state.finish()


@dp.message_handler(state=fsm.Video.link)
async def sticker_id(msg: types.Message, state: FSMContext) -> None:
    await state.update_data(link=msg.text)
    await msg.answer(f'<code>video_training</code> - инструкция от Тимура\n\n'
                     f'Укажите code для добавления видео (просто скопируй код из списка выше):')
    await fsm.Video.code.set()


@dp.message_handler(state=fsm.Video.code)
async def sticker_id(msg: types.Message, state: FSMContext) -> None:
    code = msg.text
    data = await state.get_data()
    link = data['link']
    db.add_video(code, link)
    await msg.answer('Видео добавлено!')
    await state.finish()


# @dp.message_handler(state=fsm.Check.first)
# async def text_check(msg: types.Message, state: FSMContext) -> None:
#     await send_comment_partner(msg, state)
#     await state.finish()

# Прописать очистку state и добавлять видео в БД только после подтвердить
@dp.message_handler(commands=['start'], state='*')
async def cmd_start(msg: types.Message, state: FSMContext) -> None:
    ref = msg.text[7:]
    if not db.admin_exists(msg.from_user.id):
        if ref != '':
            if db.admin_exists(int(ref)):
                db.add_admin(msg.from_user.id, msg.from_user.username, msg.from_user.first_name, msg.from_user.last_name)
            else:
                await msg.answer(db.get_msg('not_admin'))
        else:
                await msg.answer(db.get_msg('not_admin'))
    else:
        admin = db.get_admin(msg.from_user.id)
        await msg.answer(db.get_msg('main_admin'), reply_markup=kb.admin_main_menu(admin[5]))


@dp.callback_query_handler(kb.from_support_cb.filter(), state='*')
async def from_support(callback_query: types.CallbackQuery, callback_data: dict, state: FSMContext) -> None:
    msg = callback_query.message
    await bot.edit_message_reply_markup(chat_id=msg.chat.id, message_id=msg.message_id, reply_markup=None)
    await state.update_data(user_id=callback_data['user_id'])
    await bot.send_message(chat_id=callback_query.from_user.id, text=db.get_msg('from_support'))
    await fsm.ToSupport.msg.set()


@dp.message_handler(state=fsm.ToSupport.msg)
async def to_support(msg: types.Message, state: FSMContext) -> None:
    data = await state.get_data()
    user_id = data['user_id']
    text = f'Ответ от саппорта:\n\n{msg.text}'
    await user_bot.send_message(chat_id=user_id, text=text,
                                reply_markup=kb.from_support('Ответить', msg.from_user.id, 0))
    db.add_support_msg(user_id, msg.from_user.id, msg.text, 'to')
    await state.finish()


MENU = ['Добавить куратора', 'Добавить участника', 'Тексты', 'Добавить видео', 'Рассылка по группе', 'Статистика',
        'Распределить участников', 'Список участников', 'Запись практики', 'Кружок', 'Статистика по участнику']


@dp.message_handler(Text(equals=MENU), state='*')
async def admin_menu(msg: types.Message, state: FSMContext) -> None:
    item_menu = msg.text
    if item_menu == 'Добавить куратора':
        link = f'https://t.me/{cfg.ADMIN_BOT_NAME}?start={msg.from_user.id}'
        await msg.answer(f'Для добавления куратора отправьте эту ссылку на бота:\n\n<code>{link}</code>\n\n'
                         f'Скопируйте ссылку нажатием на нее')
    elif item_menu == 'Добавить участника':
        link = f'https://t.me/{cfg.USER_BOT_NAME}?start={msg.from_user.id}'
        await msg.answer(f'Для добавления нового участника отправьте эту ссылку на бота:\n\n<code>{link}</code>\n\n'
                         f'Скопируйте ссылку нажатием на нее')
    elif item_menu == 'Тексты':
        await msg.answer('В разработке')
    elif item_menu == 'Добавить видео (не для записей тренировок)':
        await msg.answer('Пришли мне ссылку на видео')
        await fsm.Video.link.set()
    elif item_menu == 'Рассылка по группе':
        await msg.answer('В разработке')
    elif item_menu == 'Запись практики':
        await msg.answer('Пришли ссылку на видео записи')
        await fsm.Record.link.set()
    elif item_menu == 'Кружок':
        await msg.answer('Пришли кружок')
        await fsm.VideoNote.video.set()
    elif item_menu == 'Статистика':
        await all_statistics(msg.from_user.id)
    elif item_menu == 'Распределить участников':
        curators = db.get_admins(status=None)
        for curator in curators:
            await bot.send_message(chat_id=curator[1], reply_markup=kb.users2curators(),
                                   text='Распределите участников. Нажмите на кнопку ниже')
    elif item_menu == 'Список участников':
        users = db.get_users_curator(msg.from_user.id)
        text = 'Список ваших участников:\n\n'
        for user in users:
            text = text + f'{user[2]}\n'
        await msg.answer(text)
    elif item_menu == 'Статистика по участнику':
        await bot.send_message(chat_id=msg.from_user.id, text='Введите username участника')
        await fsm.Statistic.username.set()
    else:
        print('ERROR admin_menu')
    await bot.delete_message(chat_id=msg.chat.id, message_id=msg.message_id)


@dp.callback_query_handler(lambda c: c.data == 'users2curators', state='*')
async def users2curators(callback_query: types.CallbackQuery, state: FSMContext) -> None:
    await bot.send_message(chat_id=callback_query.from_user.id,
                           text='Отправьте список юзернеймов пользователей под вашим кураторством')
    await fsm.Users2Curators.wait.set()
    await bot.delete_message(chat_id=callback_query.message.chat.id, message_id=callback_query.message.message_id)


@dp.message_handler(state=fsm.Record.link)
async def record_mailing(msg: types.Message, state: FSMContext) -> None:
    text = db.get_msg('training_record')
    await state.update_data(text=text)
    await state.update_data(link=msg.text)
    await msg.answer('Подтвердите отправку рассылки с ссылкой на практику', reply_markup=kb.record_confirm())
    await fsm.Record.confirm.set()


@dp.callback_query_handler(kb.record_confirm_cb.filter(), state=fsm.Record.confirm)
async def record_confirm_btn(callback_query: types.CallbackQuery, callback_data: dict, state: FSMContext) -> None:
    await bot.answer_callback_query(callback_query.id)
    msg = callback_query.message
    data = await state.get_data()
    text = data['text']
    link = data['link']
    await bot.edit_message_reply_markup(chat_id=msg.chat.id, message_id=msg.message_id, reply_markup=None)
    if callback_data['action'] == 'yes':
        date = datetime.now().strftime('%d.%m.%Y')
        db.add_training_record(link, date)
        await msg.answer('Запись добавлена в базу данных!')
        users = db.get_users()
        history = db.get_users_history(date)
        history_users = list()
        for i in history:
            history_users.append(i[0])
        for user in users:
            if user[1] not in history_users:
                await user_bot.send_message(chat_id=user[1], text=text, reply_markup=kb.skip_training(),
                                            disable_web_page_preview=True)
        await bot.send_message(chat_id=callback_query.from_user.id, text='Рассылка отправлена')
    else:
        await bot.send_message(chat_id=callback_query.from_user.id, text='Рассылка отменена')
    await state.finish()


@dp.message_handler(state=fsm.Users2Curators.wait)
async def get_users2curators(msg: types.Message, state: FSMContext) -> None:
    list_users = msg.text.split('\n')
    for user in list_users:
        username = user.replace('@', '').replace(' ', '')
        try:
            db.add_curator(username, msg.from_user.id)
        except Exception as ex:
            print('ERROR users2curators')
    await msg.answer('Готово! Список участников вы можете посмотреть в статистике')
    await state.finish()
    await bot.delete_message(chat_id=msg.chat.id, message_id=msg.message_id)


@dp.message_handler(content_types=types.ContentType.ANY, state=fsm.VideoNote.video)
async def get_video_note(msg: types.Message, state: FSMContext) -> None:
    file_id = msg.video_note.file_id
    file = await bot.get_file(file_id)
    await bot.download_file(file.file_path, 'video.mp4')
    await bot.send_video_note(chat_id=msg.from_user.id, video_note=types.InputFile('video.mp4'))
    await bot.send_message(chat_id=msg.from_user.id, text='Подтвердить рассылку кружка?',
                           reply_markup=kb.video_note_confirm())
    await fsm.VideoNote.confirm.set()


@dp.callback_query_handler(kb.video_note_confirm_cb.filter(), state=fsm.VideoNote.confirm)
async def video_note_confirm_btn(callback_query: types.CallbackQuery, callback_data: dict, state: FSMContext) -> None:
    await bot.answer_callback_query(callback_query.id)
    msg = callback_query.message
    await bot.edit_message_reply_markup(chat_id=msg.chat.id, message_id=msg.message_id, reply_markup=None)
    if callback_data['action'] == 'yes':
        users = db.get_users()
        await bot.send_message(chat_id=callback_query.from_user.id, text='Отправка началась...')
        video_note_id = ''
        for user in users:
            try:
                if video_note_id == '':
                    last_msg = await user_bot.send_video_note(chat_id=user[1], video_note=types.InputFile('video.mp4'))
                    video_note_id = last_msg.video_note.file_id
                else:
                    await user_bot.send_video_note(chat_id=user[1], video_note=video_note_id)
            except Exception as ex:
                print(ex)
        await bot.send_message(chat_id=callback_query.from_user.id, text='Рассылка отправлена')
    else:
        await bot.send_message(chat_id=callback_query.from_user.id, text='Рассылка отменена')
    await state.finish()


@dp.message_handler(content_types=types.ContentType.AUDIO, state='*')
async def get_audio(msg: types.Message, state: FSMContext) -> None:
    print(msg)


# Эхо функция
@dp.message_handler()
async def echo(msg: types.Message) -> None:
    print(msg)


async def on_startup(_):
    print('Админ-Бот запущен!')


if __name__ == "__main__":
    executor.start_polling(dispatcher=dp, skip_updates=True, on_startup=on_startup)
